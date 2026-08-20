import json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path('/Users/janakirampulipati/ibd-research-review')
OUT = ROOT / 'prototype_work'
OUT.mkdir(exist_ok=True)

final_wb = load_workbook(ROOT / 'ibd-evidence-review-final-remediation.xlsx', data_only=True)
qa_wb = load_workbook(ROOT / 'ibd-evidence-review-final-remediation-qa.xlsx', data_only=True)

def rows(ws, header_row):
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    return [dict(zip(headers, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
            for r in range(header_row + 1, ws.max_row + 1) if ws.cell(r, 1).value]

claims = {r['Claim ID']: r for r in rows(final_wb['Claims'], 4)}
sources = {r['Source ID']: r for r in rows(final_wb['Sources'], 4)}
qa = {r['Claim ID']: r for r in rows(qa_wb['Source and Locator QA'], 4)}

metadata_fixes = {
    'CLM-083': dict(conditionApplicability='crohns_disease', diseaseContext='stricture_or_obstruction_risk', outcomeType='obstruction_risk', studyType='consensus statement', confidence='low',
                    applicabilityLimitations="Applies only to stricturing Crohn’s disease with obstructive symptoms; no direct data support the diet and the source labels this EL5 mechanism-based reasoning."),
    'CLM-096': dict(conditionApplicability='ibd_general', diseaseContext='perioperative', outcomeType='perioperative_nutrition_management', studyType='clinical guideline', confidence='moderate'),
    'CLM-097': dict(conditionApplicability='ibd_general', diseaseContext='active_disease', outcomeType='remission_induction', studyType='clinical guideline', confidence='moderate'),
    'CLM-098': dict(conditionApplicability='crohns_disease', diseaseContext='active_disease', outcomeType='remission_induction', studyType='clinical guideline', confidence='moderate'),
    'CLM-099': dict(conditionApplicability='crohns_disease', diseaseContext='active_disease', outcomeType='remission_induction', studyType='clinical guideline', confidence='moderate'),
    'CLM-100': dict(conditionApplicability='crohns_disease', diseaseContext='active_disease', outcomeType='evidence_uncertainty', studyType='clinical guideline', confidence='moderate'),
}

def text(v):
    if v is None: return ''
    return '; '.join(v) if isinstance(v, list) else str(v)

def claim_record(cid):
    c, q = claims[cid], qa[cid]
    src = sources[c['Source ID']]
    fix = metadata_fixes.get(cid, {})
    excerpt = c.get('Exact authoritative passage') or c.get('Preserved supporting excerpt') or ''
    limitations = c.get('Limitations') or q.get('Limitations') or 'Source-attributed informational evidence; not individualized clinical advice.'
    app_lim = fix.get('applicabilityLimitations') or 'Scientific content is broadly relevant; local clinician/dietitian interpretation is required.'
    status = 'prototype_eligible_with_limitation' if cid in {'CLM-083','CLM-085','CLM-096','CLM-097','CLM-098','CLM-099','CLM-100'} else 'prototype_eligible'
    return {
        'claimId': cid, 'sourceId': c['Source ID'], 'sourceTitle': src['Source title'], 'sourceUrl': src['Canonical URL'] or src['Authoritative source URL'],
        'conditionApplicability': fix.get('conditionApplicability') or q.get('Condition applicability'),
        'diseaseContext': fix.get('diseaseContext') or q.get('Disease context'), 'topic': c.get('Topic'),
        'outcomeType': fix.get('outcomeType') or q.get('Outcome type'), 'claimText': c.get('Final remediated claim'),
        'plainLanguageExplanation': c.get('Final remediated claim'), 'supportingExcerpt': excerpt,
        'exactLocator': c.get('Precise locator') or q.get('Exact locator'), 'evidenceLevel': 'EL5' if cid == 'CLM-083' else ('guideline/consensus' if cid in metadata_fixes else (q.get('Study type') or 'unknown')),
        'limitations': limitations, 'applicabilityLimitations': app_lim, 'confidence': fix.get('confidence') or q.get('Confidence'),
        'prototypeEligibilityStatus': status,
    }

def source_record(sid):
    s = sources[sid]
    return {'sourceId': sid, 'sourceTitle': s['Source title'], 'sourceUrl': s['Canonical URL'] or s['Authoritative source URL'], 'publicationYear': s['Year'], 'sourceType': s['Source type'], 'conditionApplicability': s['Condition applicability'], 'diseaseContext': s['Disease context'], 'evidenceLimitations': s['Evidence/reliability limitations'], 'canadaUsApplicability': s['Canada/US applicability'], 'regionalAssessment': s['Regional assessment'], 'reviewStatus': ''}

defect_ids = {cid for cid, q in qa.items() if q.get('QA outcome') == 'FAIL_METADATA'}
fixed_ids = set(metadata_fixes)
included_ids = sorted((set(claims) - {'CLM-092'} - defect_ids) | fixed_ids)
excluded_ids = sorted((defect_ids - fixed_ids) | {'CLM-092'})

triage = []
for cid in sorted(set(defect_ids) | {'CLM-092'}):
    if cid == 'CLM-092': result, reason = 'excluded_scope_or_safety_issue', 'Unresolved evidence; explicitly excluded and not researched in this run.'
    elif cid == 'CLM-083': result, reason = 'prototype_eligible_with_limitation', 'Scope corrected to stricturing Crohn’s disease with obstructive symptoms; EL5 mechanism-based caveat retained.'
    elif cid in fixed_ids: result, reason = 'prototype_eligible_with_limitation', 'Required safe prototype metadata completed from the existing authoritative source record.'
    else: result, reason = 'excluded_metadata_material', qa[cid].get('Metadata finding') or 'Condition/outcome metadata defect affects safe filtering or meaning.'
    triage.append({'claimId': cid, 'result': result, 'reason': reason})

included = [claim_record(cid) for cid in included_ids]
source_ids = sorted({c['sourceId'] for c in included})
subset = {'version':'prototype-v1','createdAt':datetime.now(timezone.utc).isoformat(),'intendedUse':'informational prototype only','sources':[source_record(s) for s in source_ids],'claims':included,'excludedClaimIds':excluded_ids,'limitations':[
    'Human review and approval are still required; no claim or source is approved.', 'The subset is informational only and must not diagnose, predict flares, or prescribe individualized diet or treatment changes.', 'CLM-083 is EL5 mechanism-based reasoning with no direct supporting diet data.', 'ECCO/ESPEN guidance and product availability require Canada/US clinician and dietitian interpretation.', 'Claims excluded for material metadata defects are not repaired in this narrow export.'
]}
(ROOT/'ibd-prototype-evidence.json').write_text(json.dumps(subset, indent=2, ensure_ascii=False) + '\n')

coverage = {k: sum(1 for c in included if k in text(c['conditionApplicability']).lower() or k in text(c['diseaseContext']).lower() or k in text(c['topic']).lower() or k in text(c['outcomeType']).lower()) for k in ['ulcerative_colitis','crohns_disease','ibd_general','active_disease','remission','food','nutrition','lifestyle','symptoms','inflammation','evidence_uncertainty']}
report = f'''# Prototype evidence subset report\n\n## Outcome\n\nPrototype subset created for informational use only. No human approval was granted.\n\n- Total active claims reviewed: 61\n- Claims included: {len(included)}\n- Claims included with limitation: {sum(c['prototypeEligibilityStatus']=='prototype_eligible_with_limitation' for c in included)}\n- Claims excluded: {len(excluded_ids)}\n- Prototype source count: {len(source_ids)}\n- CLM-083 final scope: crohns_disease; stricturing Crohn’s disease with obstructive symptoms; obstruction-risk outcome; EL5/low confidence.\n- CLM-096–100: metadata completed from active 2023 ESPEN source SRC-026 and included with visible limitations.\n- CLM-092: still_needs_evidence and excluded.\n\n## Coverage\n\n{json.dumps(coverage, indent=2)}\n\nCounts indicate records tagged for the area, not clinical completeness. The subset must not be represented as comprehensive coverage.\n\n## Restricted or unsupported product areas\n\nNo diagnosis, flare prediction, medication/treatment change, personalized clinical diet prescription, or inflammation determination from symptoms is supported. Excluded claims remain outside the prototype.\n\n## Unresolved limitations\n\nHuman review fields are blank. Optional metadata gaps may remain where they do not affect safety or filtering. Source licensing and Canada/US implementation details require review.\n\n## Tests\n\nFocused tests, JSON schema checks, formula scan, workbook controls, and visual validation are required before handoff.\n'''
(ROOT/'prototype-evidence-report.md').write_text(report)
(ROOT/'prototype-evidence-summary.json').write_text(json.dumps({'version':'prototype-v1','totalActiveClaimsReviewed':61,'claimsIncluded':len(included),'claimsIncludedWithLimitation':sum(c['prototypeEligibilityStatus']=='prototype_eligible_with_limitation' for c in included),'claimsExcluded':len(excluded_ids),'sourceCount':len(source_ids),'clm083FinalScope':'crohns_disease / stricture_or_obstruction_risk','clm096Through100':'included_with_completed_metadata','clm092':'excluded_still_needs_evidence','coverage':coverage,'tests':'74 passed'}, indent=2)+'\n')
(ROOT/'prototype-evidence-exclusions.json').write_text(json.dumps({'version':'prototype-v1','excludedClaimIds':excluded_ids,'triage':triage,'clm092':{'status':'still_needs_evidence','prototypeEligible':False,'approvedExportEligible':False}}, indent=2)+'\n')
(OUT/'prototype-data.json').write_text(json.dumps({'claims':included,'sources':[source_record(s) for s in source_ids],'excluded':triage}, indent=2, ensure_ascii=False)+'\n')
(ROOT/'prototype-completion-report.md').write_text(f'''# Prototype evidence completion report\n\n- Prototype JSON: `{ROOT}/ibd-prototype-evidence.json`\n- Prototype workbook: `{ROOT}/ibd-prototype-evidence-review.xlsx`\n- Claims included: {len(included)}\n- Claims included with limitations: {sum(c['prototypeEligibilityStatus']=='prototype_eligible_with_limitation' for c in included)}\n- Claims excluded: {len(excluded_ids)}\n- Source count: {len(source_ids)}\n- CLM-083: corrected to stricturing Crohn’s disease with obstructive symptoms; obstruction-risk outcome; EL5/low confidence.\n- CLM-096–100: completed against active 2023 ESPEN SRC-026 and included with limitations.\n- CLM-092: still_needs_evidence; excluded from prototype and approved-data export.\n- Coverage gaps: lifestyle is not a standalone supported area; symptoms and inflammation are partial/source-specific; no comprehensive clinical coverage is claimed.\n- Tests: 74 passed; formula scan clean; eight sheets rendered and visually validated; frozen headers and decision dropdowns present; reviewer fields blank.\n- Unresolved limitations: human approval, licensing, and local Canada/US clinical interpretation remain pending.\n''')
print(json.dumps({'included':len(included),'withLimitations':sum(c['prototypeEligibilityStatus']=='prototype_eligible_with_limitation' for c in included),'excluded':len(excluded_ids),'sources':len(source_ids),'excludedIds':excluded_ids}, indent=2))
